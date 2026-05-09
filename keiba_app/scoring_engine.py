import os
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

class ScoringEngine:
    """
    Excel ファイルの読み込み・正誤判定・Excel レポート出力を担うクラス。
    UI に一切依存しないため、単体テストや CLI からも呼び出せる。

    Attributes
    ----------
    correct_map : dict[int, str | None]  正解マスタ {問番号: 答え}
    user_map    : dict[int, str | None]  解答データ {問番号: 答え}
    score       : int    正解数
    valid_count : int    有効問題数（正解が None でない問題の数）
    percentage  : float  得点率（0〜100）
    rows_data   : list   Excel 出力用の行データ
    judgments   : dict   {問番号: (is_correct, is_valid)} 判定結果
    """

    # ── ランク閾値 ───────────────────────────────────────────────────────
    RANK_THRESHOLDS = {
        "S": 100,
        "A": 70,
        "B": 50,
        "C": 0,
    }

    RESULT_MESSAGES = {
        "S": ("🌟🏆 [G1制覇] 伝説の三冠馬級！", "#FFD700"),
        "A": ("🥈 [重賞入着] 素晴らしい末脚です", "#C0C0C0"),
        "B": ("🐎 [入賞] 掲示板に載りました", "#FFCC99"),
        "C": ("🏃 [未勝利] ゲート練習からやり直し", "#A9A9A9"),
    }

    def __init__(self):
        self.correct_map: dict = {}
        self.user_map: dict = {}
        self.score: int = 0
        self.valid_count: int = 0
        self.percentage: float = 0.0
        self.rows_data: list = []
        self.judgments: dict = {}

    # ── Excel 読み込み ──────────────────────────────────────────────────
    @staticmethod
    def load_answers(file_path: str) -> dict:
        """
        Excel ファイルを読み込んで {問番号: 答え文字列} の辞書を返す。
        列レイアウト: (問番号, 答え) のペアが 4 列組（A-B, C-D, E-F, G-H）。
        """
        try:
            df = pd.read_excel(file_path, header=None)
            data_map: dict = {}
            col_pairs = [(0, 1), (2, 3), (4, 5), (6, 7)]

            # 1行目が数値でなければヘッダー行とみなしてスキップ
            start_row = 1 if not str(df.iloc[0, 0]).isdigit() else 0

            for col_num_idx, col_ans_idx in col_pairs:
                for row_idx in range(10):
                    try:
                        target_row = start_row + row_idx
                        if target_row >= len(df):
                            continue

                        q_val = df.iloc[target_row, col_num_idx]
                        ans_val = df.iloc[target_row, col_ans_idx]

                        if pd.notna(q_val):
                            q_num = int(float(q_val))
                            # float で読まれた整数（例: 3.0）の小数点以下を除去
                            s_val = (
                                str(ans_val).split(".")[0]
                                if isinstance(ans_val, float)
                                else str(ans_val)
                            )
                            data_map[q_num] = (
                                s_val.strip().upper()
                                if pd.notna(ans_val) and s_val != "nan"
                                else None
                            )
                    except Exception:
                        continue  # 読み取れないセルは無視

            return data_map

        except Exception as e:
            raise RuntimeError(f"エクセル読込エラー: {e}") from e

    # ── 採点メイン ────────────────────────────────────────────────────
    def grade(self, correct_file: str, user_file: str) -> None:
        """
        2 つの Excel ファイルを読み込み、採点結果を各 Attribute に格納する。
        """
        self.correct_map = self.load_answers(correct_file)
        self.user_map = self.load_answers(user_file)

        all_qs = sorted(self.correct_map.keys())
        self.rows_data = []
        self.judgments = {}
        self.score = 0
        self.valid_count = 0

        for q in all_qs:
            c_ans = self.correct_map[q]
            u_ans = self.user_map.get(q)
            is_valid = c_ans is not None
            is_correct = (str(u_ans) == str(c_ans)) if is_valid else False

            if is_valid:
                self.valid_count += 1
                if is_correct:
                    self.score += 1

            self.judgments[q] = (is_correct, is_valid)
            self.rows_data.append(
                [
                    q,
                    u_ans if u_ans else "未記入",
                    c_ans if is_valid else "-",
                    (
                      "⭕"
                       if (is_valid and is_correct)
                       else (
                            "✖"
                            if is_valid
                            else "-"
                        )
                    ),
                ]
            )

        self.percentage = (
            (self.score / self.valid_count * 100) if self.valid_count > 0 else 0.0
        )

    # ── ランク判定 ────────────────────────────────────────────────────
    def get_rank(self) -> str:
        """得点率からランク文字列（S/A/B/C）を返す。"""
        if self.percentage == 100:
            return "S"
        elif self.percentage >= 70:
            return "A"
        elif self.percentage >= 50:
            return "B"
        else:
            return "C"

    def get_result_message(self) -> tuple[str, str]:
        """ランクに対応する (メッセージ, カラーコード) を返す。"""
        return self.RESULT_MESSAGES[self.get_rank()]

    # ── Excel レポート出力 ────────────────────────────────────────────
    def export_excel(self, output_file: str) -> None:
        """
        採点結果を Excel ファイルに書き出す。
        左右2ブロック表示 + 中央空列付き。
        """

        all_qs = sorted(self.correct_map.keys())

        # 左右2ブロック + 中央空列
        report_df = pd.DataFrame(
            index=range(20),
            columns=[
                "問題", "解答", "正解", "判定",
                "",
                "問題", "解答", "正解", "判定"
            ],
        )

        for i in range(min(len(all_qs), 40)):

            # 左: A〜D
            # 右: F〜I
            col_offset = 0 if i < 20 else 5

            report_df.iloc[
                i % 20,
                col_offset : col_offset + 4
            ] = self.rows_data[i]

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

            report_df.to_excel(
                writer,
                sheet_name="レース結果",
                index=False
            )

            ws = writer.sheets["レース結果"]

            # スタイル適用
            self._apply_excel_styles(ws, all_qs)

            # 列幅調整
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 10
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 10

            # 中央空列
            ws.column_dimensions["E"].width = 5

            ws.column_dimensions["F"].width = 10
            ws.column_dimensions["G"].width = 10
            ws.column_dimensions["H"].width = 10
            ws.column_dimensions["I"].width = 10

    def _apply_excel_styles(self, ws, all_qs: list) -> None:
        """
        Excelスタイル適用
        """

        header_fill = PatternFill(
            start_color="DA1F28",
            end_color="DA1F28",
            fill_type="solid"
        )

        ok_fill = PatternFill(
            start_color="E6FFFA",
            end_color="E6FFFA",
            fill_type="solid"
        )

        ng_fill = PatternFill(
            start_color="FFEBEE",
            end_color="FFEBEE",
            fill_type="solid"
        )

        answer_fill = PatternFill(
            start_color="FFF8DC",
            end_color="FFF8DC",
            fill_type="solid"
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for r in range(1, 22):

            for c in range(1, 10):

                cell = ws.cell(row=r, column=c)

                # 中央空列(E列)
                if c == 5:
                    continue

                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

                # ヘッダー
                if r == 1:

                    cell.fill = header_fill

                    cell.font = Font(
                        color="FFFFFF",
                        bold=True
                    )

                else:

                    # 左右ブロック対応
                    q_idx = (
                        0 if c <= 4 else 20
                    ) + (r - 2)

                    if q_idx < len(all_qs):

                        is_ok, is_valid = self.judgments[
                            all_qs[q_idx]
                        ]

                        # 解答列(B/G)
                        if c in (2, 7):
                            cell.fill = answer_fill

                        # その他セル
                        elif is_valid:
                            cell.fill = (
                                ok_fill if is_ok else ng_fill
                            )

                        # 判定列
                        if c in (4, 9):

                            cell.font = Font(
                                color=(
                                    "0000FF"
                                    if is_ok
                                    else "FF0000"
                                ),
                                bold=True,
                                size=14
                            )